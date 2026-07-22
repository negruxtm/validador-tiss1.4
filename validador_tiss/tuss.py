"""Catálogo TUSS local e validação dos códigos de procedimentos do XML."""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from lxml import etree

from .core import ErroValidacao, _localname

APP_DIR = Path.home() / ".validador_tiss"
CATALOGO_PADRAO = APP_DIR / "tuss.json"

COLUNAS_CODIGO = ("codigo", "codigo tuss", "codigo do termo", "cd_procedimento", "cd_tuss", "cod_tuss")
COLUNAS_DESCRICAO = ("termo", "descricao do termo", "descricao", "ds_procedimento", "procedimento")
COLUNAS_INICIO = ("inicio de vigencia", "data inicio", "dt_inicio", "vigencia")
COLUNAS_FIM = ("fim de vigencia", "data fim", "dt_fim", "data de fim de implantacao")


def _normalizar_texto(valor: object) -> str:
    import unicodedata
    texto = unicodedata.normalize("NFKD", str(valor or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", " ", texto.lower()).strip()


def _normalizar_codigo(valor: object) -> str:
    texto = str(valor or "").strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return re.sub(r"\D", "", texto)


def _achar_coluna(cabecalhos: Iterable[object], candidatos: tuple[str, ...]) -> object | None:
    mapa = {_normalizar_texto(c): c for c in cabecalhos if c is not None}
    for candidato in candidatos:
        chave = _normalizar_texto(candidato)
        if chave in mapa:
            return mapa[chave]
        # Mantém a prioridade dos candidatos. Ex.: "fim de vigência" deve ser
        # preferido a "fim de implantação" mesmo quando apenas este último é exato.
        for normalizado, original in mapa.items():
            if chave in normalizado:
                return original
    return None


def _data_iso(valor: object) -> str | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, (date, datetime)):
        return valor.date().isoformat() if isinstance(valor, datetime) else valor.isoformat()
    texto = str(valor).strip()
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto[:10], formato).date().isoformat()
        except ValueError:
            pass
    return None


@dataclass(frozen=True)
class ProcedimentoTUSS:
    codigo: str
    descricao: str = ""
    inicio_vigencia: str | None = None
    fim_vigencia: str | None = None

    def vigente_em(self, referencia: date) -> bool:
        inicio = date.fromisoformat(self.inicio_vigencia) if self.inicio_vigencia else None
        fim = date.fromisoformat(self.fim_vigencia) if self.fim_vigencia else None
        return (inicio is None or referencia >= inicio) and (fim is None or referencia <= fim)


class CatalogoTUSS:
    def __init__(self, procedimentos: dict[str, ProcedimentoTUSS] | None = None, origem: str = ""):
        self.procedimentos = procedimentos or {}
        self.origem = origem

    def __len__(self) -> int:
        return len(self.procedimentos)

    def contem(self, codigo: str, referencia: date | None = None) -> bool:
        procedimento = self.procedimentos.get(_normalizar_codigo(codigo))
        return bool(procedimento and (referencia is None or procedimento.vigente_em(referencia)))

    def salvar(self, caminho: str | Path = CATALOGO_PADRAO) -> Path:
        destino = Path(caminho)
        destino.parent.mkdir(parents=True, exist_ok=True)
        dados = {"origem": self.origem, "procedimentos": [p.__dict__ for p in self.procedimentos.values()]}
        destino.write_text(json.dumps(dados, ensure_ascii=False, indent=2), encoding="utf-8")
        return destino

    @classmethod
    def carregar(cls, caminho: str | Path = CATALOGO_PADRAO) -> "CatalogoTUSS":
        arquivo = Path(caminho)
        if not arquivo.exists():
            return cls()
        dados = json.loads(arquivo.read_text(encoding="utf-8"))
        itens = {item["codigo"]: ProcedimentoTUSS(**item) for item in dados.get("procedimentos", [])}
        return cls(itens, dados.get("origem", ""))

    @classmethod
    def importar(cls, caminho: str | Path) -> "CatalogoTUSS":
        arquivo = Path(caminho)
        if arquivo.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError("Para importar Excel, instale a dependência openpyxl.") from exc
            planilha = load_workbook(arquivo, read_only=True, data_only=True)
            linhas: list[dict] = []
            for aba in planilha.worksheets:
                valores = aba.iter_rows(values_only=True)
                cabecalho = None
                # As planilhas oficiais da ANS possuem capa/título antes da tabela.
                # Procura a linha real de cabeçalho em vez de assumir a primeira linha.
                for numero_linha, candidata in enumerate(valores, start=1):
                    if _achar_coluna(candidata, COLUNAS_CODIGO) is not None:
                        cabecalho = candidata
                        break
                    if numero_linha >= 100:
                        break
                if not cabecalho:
                    continue
                for linha in valores:
                    registro = dict(zip(cabecalho, linha))
                    if any(valor not in (None, "") for valor in registro.values()):
                        linhas.append(registro)
        elif arquivo.suffix.lower() == ".csv":
            texto = arquivo.read_text(encoding="utf-8-sig", errors="replace")
            amostra = texto[:4096]
            try:
                dialeto = csv.Sniffer().sniff(amostra, delimiters=";,\t|")
            except csv.Error:
                dialeto = csv.excel
            linhas = list(csv.DictReader(texto.splitlines(), dialect=dialeto))
        else:
            raise ValueError("Formato não aceito. Selecione um arquivo CSV ou XLSX.")

        if not linhas:
            raise ValueError("A tabela selecionada não contém registros.")
        cabecalhos = linhas[0].keys()
        col_codigo = _achar_coluna(cabecalhos, COLUNAS_CODIGO)
        if not col_codigo:
            raise ValueError("Não foi encontrada uma coluna de código TUSS na tabela.")
        col_descricao = _achar_coluna(cabecalhos, COLUNAS_DESCRICAO)
        col_inicio = _achar_coluna(cabecalhos, COLUNAS_INICIO)
        col_fim = _achar_coluna(cabecalhos, COLUNAS_FIM)

        procedimentos: dict[str, ProcedimentoTUSS] = {}
        for linha in linhas:
            codigo = _normalizar_codigo(linha.get(col_codigo))
            if not codigo:
                continue
            procedimentos[codigo] = ProcedimentoTUSS(
                codigo=codigo,
                descricao=str(linha.get(col_descricao) or "").strip() if col_descricao else "",
                inicio_vigencia=_data_iso(linha.get(col_inicio)) if col_inicio else None,
                fim_vigencia=_data_iso(linha.get(col_fim)) if col_fim else None,
            )
        if not procedimentos:
            raise ValueError("Nenhum código TUSS válido foi encontrado na tabela.")
        return cls(procedimentos, arquivo.name)


ELEMENTOS_CODIGO_PROCEDIMENTO = {"codigoProcedimento", "codigoProcedimentoInterno"}


def validar_procedimentos_tuss(tree: etree._ElementTree, catalogo: CatalogoTUSS) -> list[ErroValidacao]:
    """Valida códigos quando a tabela usada no item indica TUSS (22)."""
    ocorrencias: set[tuple[str, str]] = set()
    erros: list[ErroValidacao] = []
    for elemento in tree.getroot().iter():
        if _localname(elemento.tag) not in ELEMENTOS_CODIGO_PROCEDIMENTO:
            continue
        codigo = _normalizar_codigo(elemento.text)
        if not codigo:
            continue
        pai = elemento.getparent()
        tabela = ""
        if pai is not None:
            for irmao in pai:
                if _localname(irmao.tag) == "codigoTabela":
                    tabela = (irmao.text or "").strip()
                    break
        if tabela and tabela != "22":
            continue
        chave = (codigo, tree.getpath(elemento))
        if chave in ocorrencias:
            continue
        ocorrencias.add(chave)
        if not catalogo.contem(codigo):
            erros.append(ErroValidacao(
                codigo="TUSS-NAO-ENCONTRADO",
                severidade="ERRO",
                mensagem=f"Código de procedimento {codigo} não encontrado na tabela TUSS carregada.",
                linha=elemento.sourceline,
                caminho_xml=tree.getpath(elemento),
            ))
    return erros
